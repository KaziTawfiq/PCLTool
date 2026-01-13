"""
Purpose: Serve an API that fills the grading tool .xlsm template (macros preserved) using Pole/X/Y/Z copied from BOM.
Name: main.py
Date created: 2026-01-08
Method: FastAPI endpoint receives tracker type + arrays; openpyxl loads .xlsm with keep_vba=True; writes to Inputs sheet;
        returns updated workbook bytes as download.
Data dictionary:
  Inputs:
    - tracker_type (str): "flat" or "xtr"
    - x (list): Easting values
    - y (list): Northing values
    - z (list): Elevation values
    - pole (list): Pole/description values
  Outputs:
    - .xlsm file bytes as HTTP response attachment
  Variables:
    - template_path (str): filesystem path to chosen template
    - wb: openpyxl workbook
    - ws: Inputs worksheet
"""

from fastapi import FastAPI, HTTPException
from fastapi.responses import StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel, Field
from typing import List, Any
import io
import os
from openpyxl import load_workbook

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "http://localhost:5173",
        "http://127.0.0.1:5173",
        "http://localhost:8000",
        "http://127.0.0.1:8000",
    ],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)



TEMPLATES_DIR = os.path.join(os.path.dirname(__file__), "templates")


class FillRequest(BaseModel):
    tracker_type: str = Field(..., description="flat or xtr")
    x: List[Any]
    y: List[Any]
    z: List[Any]
    pole: List[Any]

def find_inputs_sheet(wb):
    # Prefer exact name "Inputs", otherwise case-insensitive match
    for name in wb.sheetnames:
        if name == "Inputs":
            return wb[name]
    for name in wb.sheetnames:
        if name.strip().lower() == "inputs":
            return wb[name]
    return None

@app.post("/api/fill-grading-tool")
def fill_grading_tool(req: FillRequest):
    tracker = req.tracker_type.strip().lower()
    if tracker not in ("flat", "xtr"):
        raise HTTPException(status_code=400, detail="tracker_type must be 'flat' or 'xtr'")

    n = min(len(req.x), len(req.y), len(req.z), len(req.pole))
    if n <= 0:
        raise HTTPException(status_code=400, detail="No rows provided")

    template_filename = "XTR.xlsm" if tracker == "xtr" else "Flat Tracker Imperial.xlsm"
    template_path = os.path.join(TEMPLATES_DIR, template_filename)

    if not os.path.exists(template_path):
        raise HTTPException(status_code=500, detail=f"Template not found: {template_filename}")

    # Load macros-preserving workbook
    wb = load_workbook(template_path, keep_vba=True)

    ws = find_inputs_sheet(wb)
    if ws is None:
        raise HTTPException(status_code=500, detail="Could not find 'Inputs' sheet in template")

    # --- IMPORTANT: You must confirm these column positions in the template's Inputs sheet ---
    # We'll assume the Inputs sheet has header row, and these columns exist:
    # Points | Easting | Northing | Elevation | Description
    #
    # This code auto-finds column indexes by header text in row 1.
    header_row = 1
    headers = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=col).value
        if val is None:
            continue
        key = str(val).strip().lower()
        headers[key] = col

    def col_for(*names):
        for nm in names:
            if nm in headers:
                return headers[nm]
        return None

    col_points = col_for("points", "point")
    col_easting = col_for("easting", "x", "eastings")
    col_northing = col_for("northing", "y", "northings")
    col_elev = col_for("elevation", "z", "rl", "level")
    col_desc = col_for("description", "pole", "id", "name")

    missing = []
    if col_points is None: missing.append("Points")
    if col_easting is None: missing.append("Easting")
    if col_northing is None: missing.append("Northing")
    if col_elev is None: missing.append("Elevation")
    if col_desc is None: missing.append("Description")

    if missing:
        raise HTTPException(
            status_code=500,
            detail=f"Inputs sheet missing expected header(s): {', '.join(missing)}. Please check template headers."
        )

    start_row = header_row + 1

    # Optional: clear old data below header (first n rows only, or more if you want)
    # Here we clear a reasonable block (first 200000 rows is too much). We'll clear until we hit blank Points for a while.
    # For now: overwrite rows 2..(n+1) and leave the rest as-is.

    for i in range(n):
        r = start_row + i
        ws.cell(row=r, column=col_points).value = i + 1
        ws.cell(row=r, column=col_easting).value = req.x[i]
        ws.cell(row=r, column=col_northing).value = req.y[i]
        ws.cell(row=r, column=col_elev).value = req.z[i]
        ws.cell(row=r, column=col_desc).value = req.pole[i]

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)

    out_name = f"GradingTool_Filled_{tracker.upper()}.xlsm"

    return StreamingResponse(
        out,
        media_type="application/vnd.ms-excel.sheet.macroEnabled.12",
        headers={"Content-Disposition": f'attachment; filename="{out_name}"'}
    )
